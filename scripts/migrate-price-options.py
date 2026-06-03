#!/usr/bin/env python3
"""
수가표(260602) + 오픈이벤트(미끼상품) → Sanity treatment.priceOptions 마이그레이션.

규칙:
- price  = 정상가(col F)  → 취소선 표기 (수가표 헤더 명시)
- discountPrice = 홈페이지 최종금액(col H) → 실판매가
- 정상가 == 최종금액(할인율 0) 이면 price=최종금액, discountPrice 없음
- 미끼상품: price=상시가, discountPrice=이벤트가, isEvent=true → 배열 맨 앞에 prepend
- 가격은 모두 부가세 별도

dry-run:  python3 scripts/migrate-price-options.py
apply  :  python3 scripts/migrate-price-options.py --apply
"""
import sys, json, re, hashlib, urllib.request, os
import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
PRICE_XLSX = os.path.expanduser(
    "~/Downloads/포에버의원_명동_수가표_최종_260602_가격인상.xlsx"
)
EVENT_XLSX = os.path.expanduser(
    "~/Downloads/포에버의원_명동_오픈이벤트_260603_가격인상분 반영(명동 미끼상품).xlsx"
)

# 세부카테고리(col B) → treatment slug
SUBCAT_TO_SLUG = {
    "울쎄라피 프라임": "ultherapy-prime",
    "써마지 FLX": "thermage-flx",
    "세르프": "serf",
    "티타늄": "titanium",
    "온다": "onda-lifting",
    "올타이트": "alltight",
    "포텐자": "potenza",
    "실펌X": "sylfirm-x",
    "인모드": "inmode",
    "슈링크 유니버스": "shrink-universe",
    "CO2": "co2-laser",
    "피코토닝(피코K)": "pico-toning",
    "루카스토닝": "lucas-toning",
    "더마브이": "derma-v",
    "악센토": "accento",
    "하이쿡스": "haecox",
    "고우리(GOURI)": "gouri",
    "리투오": "retuo",
    "래디어스": "radiesse",
    "스컬트라": "sculptra",
    "쥬베룩": "juvelook",
    "리쥬란": "rejuran",
    "메타셀(줄기세포)": "metacell-stem-cell",
    "에피티콘실리프팅": "thread-lifting",
    "잼버실리프팅": "thread-lifting",
    "주름보톡스": "botox",
    "사각턱 보톡스": "botox",
    "스킨보톡스": "skin-botox",
    "승모근 보톡스": "botox",
    "침샘보톡스": "botox",
    "다한증보톡스": "botox",
    "국산필러": "filler",
    "수입필러": "filler",
    "윤곽주사": "fat-dissolving",
    "인텐스울트라": "intense-ultra",
    "하이드라페이셜": "hydrafacial",
    "LDM": "ldm",
    "여성 제모(젠틀맥스 프로 플러스)": "gentlemax-pro-plus",
    "남성 제모(젠틀맥스 프로 플러스)": "gentlemax-pro-plus",
    "남성 무제한 제모": "gentlemax-pro-plus",
    "여성 무제한 제모": "gentlemax-pro-plus",
    # 물광주사 / 마운자로 / 위고비 → 미등록 시술, 보류
}

# 부위 그룹(area) 사용 시술 + 세부카테고리 → 그룹 라벨
AREA_GROUPS = {
    "botox": {
        "주름보톡스": "주름",
        "사각턱 보톡스": "사각턱",
        "승모근 보톡스": "승모근·종아리",
        "침샘보톡스": "침샘",
        "다한증보톡스": "다한증",
    },
    "filler": {"국산필러": "국산 필러", "수입필러": "수입 필러"},
    "gentlemax-pro-plus": {
        "여성 제모(젠틀맥스 프로 플러스)": "여성",
        "남성 제모(젠틀맥스 프로 플러스)": "남성",
        "여성 무제한 제모": "여성 무제한",
        "남성 무제한 제모": "남성 무제한",
    },
}

# 미끼상품 시술명 → slug
EVENT_TO_SLUG = {
    "리쥬란힐러": "rejuran",
    "리쥬란HB": "rejuran",
    "포텐자+쥬베룩스킨": "potenza",
    "쥬베룩볼륨": "juvelook",
    "고우리": "gouri",
    "온다": "onda-lifting",
    "바디온다": "onda-lifting",
    "인모드": "inmode",
    "슈링크": "shrink-universe",
    "써마지": "thermage-flx",
    "울쎄라": "ultherapy-prime",
    "입술필러(리쥬비엘)": "filler",
    "입술필러(레스틸렌키스)": "filler",
    "실리프팅": "thread-lifting",
    "다한증보톡스(국산)": "botox",
    "승모근보톡스(국산)": "botox",
    "주름보톡스(국산)": "botox",
    "CO2 점제거": "co2-laser",
}


def s(v):
    return "" if v is None else str(v).strip()


def to_int(v):
    if v is None or v == "":
        return None
    try:
        return int(round(float(v)))
    except (ValueError, TypeError):
        return None


def key_for(slug, idx, kind="opt"):
    h = hashlib.md5(f"{slug}-{kind}-{idx}".encode()).hexdigest()[:10]
    return f"po{h}"


def loc(text):
    return {"ko": text} if text else None


def build():
    wb = openpyxl.load_workbook(PRICE_XLSX, data_only=True)
    ws = wb.active
    # slug -> list of raw rows
    by_slug = {}
    for row in ws.iter_rows(min_row=2, values_only=True):  # row 1 = 헤더(셀 내 줄바꿈 포함)
        major, subcat, proc, area_col, dose, regular, _disc, final = (
            list(row) + [None] * 8
        )[:8]
        subcat = s(subcat)
        proc = s(proc)
        major = s(major)
        slug = SUBCAT_TO_SLUG.get(subcat)
        # 스킨케어 빈 세부카테고리(아쿠아필/모델링팩) → aqua-peel
        if not slug and major == "스킨케어" and ("아쿠아필" in proc or "모델링" in proc):
            slug = "aqua-peel"
        if not slug:
            continue
        regular_i = to_int(regular)
        final_i = to_int(final)
        if regular_i is None and final_i is None:
            continue
        by_slug.setdefault(slug, []).append(
            {
                "subcat": subcat,
                "proc": proc,
                "area_col": s(area_col),
                "dose": s(dose),
                "regular": regular_i,
                "final": final_i,
            }
        )

    # 시술별 priceOptions 구성
    result = {}
    for slug, rows in by_slug.items():
        grouped = slug in AREA_GROUPS
        opts = []
        for i, r in enumerate(rows):
            # area
            if grouped:
                area = AREA_GROUPS[slug].get(r["subcat"], "")
            else:
                area = ""
            # name = 시술명
            name = r["proc"]
            # caption = 부위/용량 (시술명에 없는 정보만)
            cap_parts = []
            if r["area_col"] and r["area_col"] not in name and r["area_col"] != area:
                cap_parts.append(r["area_col"])
            if r["dose"]:
                cap_parts.append(r["dose"])
            caption = " / ".join(cap_parts)
            # price / discount
            reg, fin = r["regular"], r["final"]
            if reg and fin and reg > fin:
                price, discount = reg, fin
            elif fin:
                price, discount = fin, None
            else:
                price, discount = reg, None
            opt = {
                "_key": key_for(slug, i),
                "_type": "priceOption",
                "name": loc(name),
                "isEvent": False,
            }
            if caption:
                opt["caption"] = loc(caption)
            if area:
                opt["area"] = area
            if price is not None:
                opt["price"] = price
            if discount is not None:
                opt["discountPrice"] = discount
            opts.append(opt)
        # area 그룹 정렬: 그룹 라벨 등장 순서 유지 (이미 행 순서 유지됨)
        result[slug] = opts

    # 미끼상품(이벤트) → prepend
    ewb = openpyxl.load_workbook(EVENT_XLSX, data_only=True)
    ews = ewb.active
    ev_by_slug = {}
    rows = list(ews.iter_rows(values_only=True))
    for row in rows[3:]:  # 헤더 3행 스킵
        gubun, proc, detail, dose, normal, event_p, _rate = (list(row) + [None] * 7)[:7]
        proc = s(proc)
        detail = s(detail)
        dose = s(dose)
        if not proc:
            continue
        slug = EVENT_TO_SLUG.get(proc)
        if not slug:
            continue
        normal_i, event_i = to_int(normal), to_int(event_p)
        if event_i is None:
            continue
        # 라벨
        if slug == "gentlemax-pro-plus":
            name = f"{detail} {proc}".strip()  # 여자 인중
        else:
            name = proc
            if detail and detail not in name:
                name = f"{name} ({detail})"
        ev_by_slug.setdefault(slug, []).append(
            {"name": name, "dose": dose, "normal": normal_i, "event": event_i}
        )

    for slug, evs in ev_by_slug.items():
        ev_opts = []
        for j, e in enumerate(evs):
            opt = {
                "_key": key_for(slug, j, "ev"),
                "_type": "priceOption",
                "name": loc(e["name"]),
                "isEvent": True,
            }
            if e["dose"]:
                opt["caption"] = loc(e["dose"])
            # price=상시가(취소선), discount=이벤트가
            if e["normal"] and e["event"] and e["normal"] > e["event"]:
                opt["price"] = e["normal"]
                opt["discountPrice"] = e["event"]
            else:
                opt["price"] = e["event"]
            ev_opts.append(opt)
        result[slug] = ev_opts + result.get(slug, [])

    return result


def fetch_token():
    env = open(os.path.join(ROOT, ".env.local")).read()
    m = re.search(r"SANITY_API_TOKEN=(.+)", env)
    return m.group(1).strip()


def fetch_id_map(token):
    q = urllib.parse.quote('*[_type=="treatment"]{"slug":slug.current,_id}')
    url = f"https://ecoamz42.api.sanity.io/v2024-01-01/data/query/production?query={q}"
    req = urllib.request.Request(url, headers={"Authorization": f"Bearer {token}"})
    data = json.loads(urllib.request.urlopen(req).read())
    return {r["slug"]: r["_id"] for r in data["result"] if r.get("slug")}


def apply(result, token, id_map):
    mutations = []
    for slug, opts in result.items():
        _id = id_map.get(slug)
        if not _id:
            print(f"  ⚠️  slug 미존재, 건너뜀: {slug}")
            continue
        mutations.append({"patch": {"id": _id, "set": {"priceOptions": opts}}})
    body = json.dumps({"mutations": mutations}).encode()
    url = "https://ecoamz42.api.sanity.io/v2024-01-01/data/mutate/production"
    req = urllib.request.Request(
        url,
        data=body,
        headers={
            "Authorization": f"Bearer {token}",
            "Content-Type": "application/json",
        },
        method="POST",
    )
    resp = json.loads(urllib.request.urlopen(req).read())
    print("  ✅ 반영 완료:", resp.get("transactionId"), "/ 문서", len(mutations), "건")


def main():
    result = build()
    total = sum(len(v) for v in result.values())
    print(f"=== 빌드 결과: {len(result)}개 시술, 총 {total}개 옵션 ===\n")
    for slug in sorted(result):
        opts = result[slug]
        ev = sum(1 for o in opts if o.get("isEvent"))
        print(f"■ {slug}  ({len(opts)}옵션, 이벤트 {ev})")
        for o in opts:
            badge = "[E]" if o.get("isEvent") else "   "
            area = f"<{o['area']}> " if o.get("area") else ""
            cap = f"  ({o['caption']['ko']})" if o.get("caption") else ""
            p = o.get("price")
            d = o.get("discountPrice")
            price = f"{d:,} (정가 {p:,})" if d else (f"{p:,}" if p else "-")
            print(f"    {badge} {area}{o['name']['ko']}{cap} → {price}")
        print()

    if "--apply" in sys.argv:
        token = fetch_token()
        id_map = fetch_id_map(token)
        print("=== Sanity production 반영 ===")
        apply(result, token, id_map)
    else:
        with open(os.path.join(HERE, "price-options.generated.json"), "w") as f:
            json.dump(result, f, ensure_ascii=False, indent=2)
        print("dry-run. --apply 로 production 반영. (JSON: scripts/price-options.generated.json)")


if __name__ == "__main__":
    main()
